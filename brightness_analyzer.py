import os
import re
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
from collections import defaultdict
from io import BytesIO
import cv2
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

SN_PATTERN = re.compile(r'^\d{10,15}$')
IMAGE_PATTERN = re.compile(r'^(left|right)(?:_e(\d+))?_p(\d+)\.png$', re.IGNORECASE)
OVEREXPOSURE_THRESHOLD = 255

def is_sn_folder(folder_name):
    return bool(SN_PATTERN.match(folder_name))

def parse_image_filename(filename):
    match = IMAGE_PATTERN.match(filename)
    if match:
        group_type = match.group(1).lower()
        expo_str = match.group(2)
        idx_str = match.group(3)
        expo = int(expo_str) if expo_str else None
        if len(idx_str) == 4:
            power_value = -1
        else:
            power_value = int(idx_str)
        return group_type, power_value, expo
    return None, None, None

def scan_folder(root_path):
    sn_folders = []
    root_path = os.path.normpath(root_path)
    root_name = os.path.basename(root_path)
    
    if is_sn_folder(root_name):
        sn_folders.append(root_path)
    else:
        for item in os.listdir(root_path):
            item_path = os.path.normpath(os.path.join(root_path, item))
            if os.path.isdir(item_path) and is_sn_folder(item):
                sn_folders.append(item_path)
    
    return sn_folders

def collect_images(folder_path):
    images = defaultdict(lambda: defaultdict(dict))
    folder_path = os.path.normpath(folder_path)
    
    for filename in os.listdir(folder_path):
        group_type, power_value, expo = parse_image_filename(filename)
        if group_type and power_value is not None:
            file_path = os.path.normpath(os.path.join(folder_path, filename))
            images[group_type][power_value] = file_path
    
    for group_type in images:
        images[group_type] = dict(sorted(images[group_type].items()))
    
    return dict(images)

def read_image_chinese_path(image_path, flags=cv2.IMREAD_COLOR):
    img_array = np.fromfile(image_path, dtype=np.uint8)
    img = cv2.imdecode(img_array, flags)
    return img

def read_image_auto_depth(image_path):
    img = read_image_chinese_path(image_path, cv2.IMREAD_UNCHANGED)
    if img is None:
        return None
    
    if len(img.shape) == 3:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    return img

def select_roi(image_path, group_type):
    image_path = os.path.normpath(image_path)
    img = read_image_chinese_path(image_path, cv2.IMREAD_COLOR)
    if img is None:
        print(f"无法读取图像: {image_path}")
        return None
    
    print(f"  操作提示: 用鼠标拖拽框选ROI区域，按 Enter 确认，按 ESC 或 C 取消")
    sys.stdout.flush()
    
    window_name = f"选择ROI - {group_type} (按Enter确认, ESC取消)"
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
    cv2.resizeWindow(window_name, 800, 600)
    
    roi = cv2.selectROI(window_name, img, showCrosshair=True, fromCenter=False)
    cv2.destroyAllWindows()
    
    if roi[2] == 0 or roi[3] == 0:
        return None
    
    return roi

def calculate_brightness_and_overexposure(image_path, roi):
    image_path = os.path.normpath(image_path)
    img = read_image_auto_depth(image_path)
    if img is None:
        return None, None
    
    x, y, w, h = roi
    roi_region = img[y:y+h, x:x+w]
    
    mean_brightness = np.mean(roi_region)
    
    if img.dtype == np.uint16:
        threshold = OVEREXPOSURE_THRESHOLD * 256
    else:
        threshold = OVEREXPOSURE_THRESHOLD
    
    overexposed_pixels = np.sum(roi_region >= threshold)
    total_pixels = roi_region.size
    overexposure_ratio = overexposed_pixels / total_pixels
    
    return mean_brightness, overexposure_ratio

def process_group(images_dict, roi_dict, group_type):
    results = []
    
    if group_type not in images_dict or group_type not in roi_dict:
        return results
    
    roi = roi_dict[group_type]
    power_images = images_dict[group_type]
    
    for power_value, image_path in power_images.items():
        brightness, overexposure_ratio = calculate_brightness_and_overexposure(image_path, roi)
        if brightness is not None:
            results.append({
                'Power': power_value,
                'Brightness': brightness,
                'OverexposureRatio': overexposure_ratio,
                'ImagePath': image_path
            })
    
    results.sort(key=lambda x: x['Power'])
    return results

def create_scatter_chart(ws, data_start_row, data_end_row, power_col, value_cols, sn_names, title, y_title):
    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = "Laser Power Value"
    chart.y_axis.title = y_title
    chart.style = 10
    chart.width = 15
    chart.height = 10
    
    x_values = Reference(ws, min_col=power_col, min_row=data_start_row, max_row=data_end_row)
    
    markers = ['circle', 'square', 'triangle', 'diamond', 'star', 'x', 'plus']
    colors = ['4472C4', 'ED7D31', 'A5A5A5', 'FFC000', '5B9BD5', '70AD47', 'FF6B6B']
    
    for idx, (col, sn_name) in enumerate(zip(value_cols, sn_names)):
        y_values = Reference(ws, min_col=col, min_row=data_start_row, max_row=data_end_row)
        series = Series(y_values, x_values, title=sn_name)
        
        marker = Marker(symbol=markers[idx % len(markers)], size=7)
        series.marker = marker
        series.graphicalProperties.line.width = 20000
        
        chart.series.append(series)
    
    return chart

def save_results_to_excel(all_results, output_path):
    output_path = os.path.normpath(output_path)
    wb = Workbook()
    
    grouped_by_type = defaultdict(dict)
    for sn_folder, group_results in all_results.items():
        sn_name = os.path.basename(sn_folder)
        for group_type, results in group_results.items():
            grouped_by_type[group_type][sn_name] = results
    
    ws_summary = wb.active
    ws_summary.title = '结果汇总'
    
    summary_row = 1
    for group_type in sorted(grouped_by_type.keys()):
        results_by_sn = grouped_by_type[group_type]
        if not results_by_sn:
            continue
        
        sn_names = sorted(results_by_sn.keys())
        
        all_powers = set()
        for sn_name, results in results_by_sn.items():
            for r in results:
                all_powers.add(r['Power'])
        all_powers = sorted(all_powers)
        
        data_start_row = summary_row + 1
        data_end_row = data_start_row + len(all_powers)
        
        headers = ['Power']
        for sn_name in sn_names:
            headers.append(f'BR【{sn_name}】')
        headers.append('')
        for sn_name in sn_names:
            headers.append(f'ORP【{sn_name}】')
        
        for col, header in enumerate(headers, 1):
            ws_summary.cell(row=summary_row, column=col, value=header)
        
        for row_idx, power in enumerate(all_powers, summary_row + 1):
            ws_summary.cell(row=row_idx, column=1, value=power)
            
            for col_idx, sn_name in enumerate(sn_names, 2):
                results = results_by_sn[sn_name]
                brightness = None
                overexposure_ratio = None
                for r in results:
                    if r['Power'] == power:
                        brightness = round(r['Brightness'], 2)
                        overexposure_ratio = round(r['OverexposureRatio'] * 100, 4)
                        break
                if brightness is not None:
                    ws_summary.cell(row=row_idx, column=col_idx, value=brightness)
                
                orp_col_idx = col_idx + len(sn_names) + 1
                if overexposure_ratio is not None:
                    ws_summary.cell(row=row_idx, column=orp_col_idx, value=overexposure_ratio)
        
        br_cols = list(range(2, 2 + len(sn_names)))
        chart1 = create_scatter_chart(ws_summary, data_start_row, data_end_row, 1, br_cols, sn_names, 
                                       f'{group_type} - 亮度分析结果', '亮度均值')
        chart1_col = get_column_letter(2 + len(sn_names) + 2)
        ws_summary.add_chart(chart1, f'{chart1_col}{summary_row}')
        
        orp_cols = list(range(2 + len(sn_names) + 1, 2 + 2 * len(sn_names) + 1))
        chart2 = create_scatter_chart(ws_summary, data_start_row, data_end_row, 1, orp_cols, sn_names,
                                       f'{group_type} - 可靠性评估', '过曝比例 (%)')
        chart2_col = get_column_letter(2 + len(sn_names) + 12)
        ws_summary.add_chart(chart2, f'{chart2_col}{summary_row}')
        
        summary_row = data_end_row + 3
    
    for group_type in sorted(grouped_by_type.keys()):
        results_by_sn = grouped_by_type[group_type]
        if not results_by_sn:
            continue
        
        ws = wb.create_sheet(title=group_type)
        
        all_powers = set()
        for sn_name, results in results_by_sn.items():
            for r in results:
                all_powers.add(r['Power'])
        all_powers = sorted(all_powers)
        
        sn_names = sorted(results_by_sn.keys())
        
        headers = ['Power']
        for sn_name in sn_names:
            headers.append(f'BR【{sn_name}】')
        headers.append('')
        for sn_name in sn_names:
            headers.append(f'ORP【{sn_name}】')
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row_idx, power in enumerate(all_powers, 2):
            ws.cell(row=row_idx, column=1, value=power)
            
            for col_idx, sn_name in enumerate(sn_names, 2):
                results = results_by_sn[sn_name]
                brightness = None
                overexposure_ratio = None
                for r in results:
                    if r['Power'] == power:
                        brightness = round(r['Brightness'], 2)
                        overexposure_ratio = round(r['OverexposureRatio'] * 100, 4)
                        break
                if brightness is not None:
                    ws.cell(row=row_idx, column=col_idx, value=brightness)
                
                orp_col_idx = col_idx + len(sn_names) + 1
                if overexposure_ratio is not None:
                    ws.cell(row=row_idx, column=orp_col_idx, value=overexposure_ratio)
    
    wb.save(output_path)
    print(f"结果已保存到: {output_path}")

def main():
    root = tk.Tk()
    root.withdraw()
    
    print("请选择要分析的文件夹...")
    selected_path = filedialog.askdirectory(title="选择分析文件夹")
    
    if not selected_path:
        print("未选择文件夹，程序退出")
        return
    
    print(f"选择的文件夹: {selected_path}")
    
    sn_folders = scan_folder(selected_path)
    
    if not sn_folders:
        messagebox.showerror("错误", "未找到符合SN命名规则的文件夹")
        return
    
    print(f"找到 {len(sn_folders)} 个SN文件夹:")
    for folder in sn_folders:
        print(f"  - {os.path.basename(folder)}")
    
    all_results = {}
    roi_dict = {}
    
    first_folder = sn_folders[0]
    first_images = collect_images(first_folder)
    
    if not first_images:
        messagebox.showerror("错误", "未找到符合条件的图像文件")
        return
    
    print("\n开始选择ROI区域...")
    for group_type in sorted(first_images.keys()):
        power_images = first_images[group_type]
        
        if 100 in power_images:
            ref_image = power_images[100]
        else:
            valid_powers = [p for p in power_images.keys() if p > 0]
            if valid_powers:
                max_power = max(valid_powers)
                ref_image = power_images[max_power]
                print(f"未找到power=100的图像，使用power={max_power}的图像作为参考")
            else:
                ref_image = list(power_images.values())[0]
                print(f"未找到有效power图像，使用第一张图像作为参考")
        
        print(f"\n请为 {group_type} 组选择ROI区域...")
        roi = select_roi(ref_image, group_type)
        
        if roi is None:
            print(f"{group_type} 组ROI选择取消，跳过该组")
            continue
        
        roi_dict[group_type] = roi
        print(f"{group_type} 组ROI: x={roi[0]}, y={roi[1]}, w={roi[2]}, h={roi[3]}")
    
    if not roi_dict:
        messagebox.showerror("错误", "未选择任何ROI区域")
        return
    
    print("\n开始处理所有SN文件夹...")
    for sn_folder in sn_folders:
        sn_name = os.path.basename(sn_folder)
        print(f"\n处理: {sn_name}")
        
        images = collect_images(sn_folder)
        if not images:
            print(f"  未找到图像文件，跳过")
            continue
        
        group_results = {}
        for group_type in roi_dict.keys():
            if group_type in images:
                print(f"  处理 {group_type} 组...")
                results = process_group(images, roi_dict, group_type)
                group_results[group_type] = results
                print(f"    完成 {len(results)} 张图像")
        
        if group_results:
            all_results[sn_folder] = group_results
    
    if not all_results:
        messagebox.showerror("错误", "未处理任何有效数据")
        return
    
    output_path = os.path.join(selected_path, 'results.xlsx')
    save_results_to_excel(all_results, output_path)
    
    messagebox.showinfo("完成", f"分析完成!\n结果已保存到:\n{output_path}")

if __name__ == '__main__':
    main()